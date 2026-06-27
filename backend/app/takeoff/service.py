"""Quantity-takeoff workbook — turn a generated test-fit into a multi-sheet bill of materials.

A Qbiq-style deliverable: the catalog-backed scene is the source of truth, so every priced
line resolves against a real `Product` row and is flagged `real` (parsed price book) vs `est.`
(synthetic fallback). Derived shell figures (areas, perimeters) are computed from the plan
geometry; partition walls, doors, and glazing are NOT yet modeled and are intentionally
omitted rather than fabricated.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet
from shapely.geometry import Polygon
from sqlalchemy.orm import Session

from ..floorplan.dxf_ingest import PlanModel
from ..models import Product
from ..testfit.bom import instance_skus
from ..testfit.layout import TestFit

ROOM_TYPE_LABELS = {
    "workstation": "Open Workstation",
    "private_office": "Private Office",
    "meeting_room": "Meeting Room",
    "collaboration": "Collaboration",
}

_HEADER_FONT = Font(bold=True)


def _header(ws: Worksheet, columns: list[str]) -> None:
    ws.append(columns)
    for cell in ws[1]:
        cell.font = _HEADER_FONT


def _inventory_rows(db: Session, fit: TestFit) -> list[dict]:
    """One row per resolved SKU, per placed instance. Missing SKUs are skipped gracefully."""
    rows: list[dict] = []
    for room_id, inst in enumerate(fit.instances):
        for sku, qty in instance_skus(inst):
            product = db.query(Product).filter(Product.sku == sku).first()
            if product is None:
                continue
            is_real = product.source == "pricebook"
            rows.append({
                "room_id": room_id,
                "room_type": ROOM_TYPE_LABELS.get(inst.type, inst.type),
                "item": product.name,
                "supplier": product.manufacturer_code,
                "qty": qty,
                "unit_price": product.list_price,
                "total": product.list_price * qty,
                "pricing": "real" if is_real else "est.",
            })
    return rows


def _furniture_inventory_sheet(ws: Worksheet, rows: list[dict]) -> None:
    _header(ws, [
        "Room ID", "Room Type", "Item", "Supplier", "Quantity",
        "Unit Price (INR)", "Total (INR)", "Pricing",
    ])
    for r in rows:
        ws.append([
            r["room_id"], r["room_type"], r["item"], r["supplier"], r["qty"],
            r["unit_price"], r["total"], r["pricing"],
        ])


def _summary_sheet(ws: Worksheet, fit: TestFit, plan: PlanModel, rows: list[dict]) -> None:
    seats = fit.workstation_count + fit.office_count
    density = plan.usable_area_sf / seats if seats else None
    grand_total = sum(r["total"] for r in rows)
    real_lines = sum(1 for r in rows if r["pricing"] == "real")
    est_lines = sum(1 for r in rows if r["pricing"] == "est.")

    _header(ws, ["Metric", "Value"])
    for label, value in [
        ("Workstations", fit.workstation_count),
        ("Private offices", fit.office_count),
        ("Meeting rooms", fit.meeting_count),
        ("Collaboration zones", fit.collab_count),
        ("Usable area (sf)", round(plan.usable_area_sf, 1)),
        ("Gross area (sf)", round(plan.gross_area_sf, 1)),
        ("Density (usable sf / seat)", round(density, 1) if density is not None else "n/a"),
        ("Grand total (INR)", round(grand_total, 2)),
        ("Real-price lines", real_lines),
        ("Estimated lines", est_lines),
    ]:
        ws.append([label, value])


def _areas_shell_sheet(ws: Worksheet, plan: PlanModel) -> None:
    boundary_perimeter = Polygon(plan.boundary).length if plan.boundary else 0.0
    core_perimeter = sum(Polygon(c).length for c in plan.cores if len(c) >= 3)

    _header(ws, ["Measure", "Value (derived)"])
    for label, value in [
        ("Gross area (sf)", round(plan.gross_area_sf, 1)),
        ("Usable area (sf)", round(plan.usable_area_sf, 1)),
        ("Boundary perimeter (LF)", round(boundary_perimeter, 1)),
        ("Total core perimeter (LF)", round(core_perimeter, 1)),
    ]:
        ws.append([label, value])
    ws.append([])
    ws.append([
        "Note",
        "Interior partition walls, doors, and glazing are NOT yet modeled and are "
        "intentionally omitted (not fabricated).",
    ])


def build_takeoff_workbook(db: Session, plan: PlanModel, fit: TestFit) -> Workbook:
    """Build the multi-sheet quantity-takeoff workbook. Pure: only reads the DB session."""
    rows = _inventory_rows(db, fit)

    wb = Workbook()
    _furniture_inventory_sheet(wb.active, rows)
    wb.active.title = "Furniture Inventory"
    _summary_sheet(wb.create_sheet("Summary"), fit, plan, rows)
    _areas_shell_sheet(wb.create_sheet("Areas & Shell"), plan)
    return wb
