"""Qbiq-grade quantity takeoff from a real ExtractedLayout — a multi-sheet .xlsx workbook.

Every QUANTITY here is computed from the layout geometry (segment lengths, panel spans, door
counts, the bounds box). Material NAMES (floor finish, ceiling, partition heights) are standard
fit-out ASSUMPTIONS, flagged "(assumed)" so a reader never mistakes them for measured values.

Lengths are in feet (the ExtractedLayout contract); areas in square feet; the Inventory sheet also
reports area in m2 for parity with the Qbiq export. Item sizes render in CM (feet x 30.48).
"""

from __future__ import annotations

import math
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from shapely.geometry import Point, Polygon

from ..ingestion.schema import ExtractedLayout, FurnitureItem
from .service import _HEADER_FONT, _header

_FT_TO_CM = 30.48
_SQFT_TO_SQM = 0.092903

# Category -> human base name for the Item Description.
_ITEM_BASE = {
    "chair": "Chairs",
    "workstation": "Workstation Table",
    "desk": "Desk Table",
    "table": "Table",
    "sofa": "Sofa Lounge",
    "stool": "Bar Chairs",
    "storage": "Storage",
    "tv": "Screen",
    "planter": "Plants",
    "panel": "Glass Partition",
}

# Room type -> (Space Type, Subcategory) for the Inventory sheet.
_SPACE_TYPE = {
    "office": ("Office", "Private Office"),
    "meeting": ("Meeting", "Conference Room"),
    "huddle": ("Meeting", "Huddle Room"),
    "open": ("Open", "Open Area"),
    "reception": ("Reception", "Reception"),
    "circulation": ("Circulation", "Circulation"),
    "core": ("Core", "Core"),
}

_CUSTOM = "Can be customized"


def _cm(feet: float) -> int:
    return round(feet * _FT_TO_CM)


def _item_description(item: FurnitureItem) -> str:
    base = _ITEM_BASE.get(item.category, item.category.title())
    return f"{base} W{_cm(item.w)} X L{_cm(item.h)}"


def _segment_length(a: tuple[float, float], b: tuple[float, float]) -> float:
    return math.hypot(b[0] - a[0], b[1] - a[1])


def _wall_length_by_type(layout: ExtractedLayout) -> dict[str, float]:
    totals: dict[str, float] = defaultdict(float)
    for wall in layout.walls:
        length = sum(
            _segment_length(wall.points[i], wall.points[i + 1])
            for i in range(len(wall.points) - 1)
        )
        totals[wall.type] += length
    return dict(totals)


def _glass_length(layout: ExtractedLayout) -> float:
    return sum(max(f.w, f.h) for f in layout.furniture if f.category == "panel")


def _box_area_sqft(layout: ExtractedLayout) -> float:
    minx, miny, maxx, maxy = layout.bounds
    return (maxx - minx) * (maxy - miny)


def _takeoff_furniture(layout: ExtractedLayout) -> list[FurnitureItem]:
    """Mullions are glazing framing, not a placed product — they never appear in the takeoff."""
    return [f for f in layout.furniture if f.category != "mullion"]


def _assign_rooms(layout: ExtractedLayout) -> tuple[dict[int, str], bool]:
    """Map each takeoff-furniture index to a Room ID by point-in-polygon on the item's CENTER.
    Returns (index -> room id, most_rooms_unpolygoned). Items in no room map to "—"."""
    polys = [
        (room.id, Polygon(room.polygon))
        for room in layout.rooms
        if len(room.polygon) >= 3
    ]
    most_unpolygoned = bool(layout.rooms) and len(polys) < len(layout.rooms) / 2

    assignment: dict[int, str] = {}
    for idx, f in enumerate(_takeoff_furniture(layout)):
        center = Point(f.x + f.w / 2.0, f.y + f.h / 2.0)
        room_id = "—"
        for rid, poly in polys:
            if poly.contains(center):
                room_id = rid
                break
        assignment[idx] = room_id
    return assignment, most_unpolygoned


def _furniture_inventory_sheet(
    ws: Worksheet, layout: ExtractedLayout, assignment: dict[int, str]
) -> None:
    _header(ws, [
        "Cost Code", "Floor", "Room ID", "Room Type", "Item Description",
        "Supplier", "Quantity", "Unit Price", "Total Price",
    ])
    rooms_by_id = {r.id: r for r in layout.rooms}
    for idx, f in enumerate(_takeoff_furniture(layout)):
        room_id = assignment[idx]
        room = rooms_by_id.get(room_id)
        room_type = room.type if room else "Unassigned"
        ws.append([
            _CUSTOM if idx == 0 else "",
            "Floor 1",
            room_id,
            room_type,
            _item_description(f),
            _CUSTOM if idx == 0 else "",
            1, 0, 0,
        ])


def _furniture_summary_sheet(ws: Worksheet, layout: ExtractedLayout) -> None:
    _header(ws, [
        "Cost Code", "Item Description", "Supplier", "Quantity", "Unit Price", "Total Price",
    ])
    counts: dict[str, int] = defaultdict(int)
    for f in _takeoff_furniture(layout):
        counts[_item_description(f)] += 1
    for i, desc in enumerate(sorted(counts)):
        ws.append([
            _CUSTOM if i == 0 else "",
            desc,
            _CUSTOM if i == 0 else "",
            counts[desc], 0, 0,
        ])


def _furniture_elements(items: list[FurnitureItem]) -> str:
    counts: dict[str, int] = defaultdict(int)
    for f in items:
        counts[_ITEM_BASE.get(f.category, f.category.title())] += 1
    return ", ".join(f"{name}: {counts[name]}" for name in sorted(counts))


def _inventory_sheet(ws: Worksheet, layout: ExtractedLayout, assignment: dict[int, str]) -> None:
    _header(ws, [
        "Floor", "Department", "Space Type", "Subcategory", "Room ID", "Program Room Name",
        "Headcount", "Area (m2)", "Area (sqf)", "Floor Material", "Ceiling Material",
        "Furniture Elements",
    ])
    takeoff = _takeoff_furniture(layout)
    by_room: dict[str, list[FurnitureItem]] = defaultdict(list)
    for idx, f in enumerate(takeoff):
        by_room[assignment[idx]].append(f)

    for room in layout.rooms:
        space_type, subcategory = _SPACE_TYPE.get(room.type, (room.type.title(), room.type.title()))
        items = by_room.get(room.id, [])
        headcount = sum(1 for f in items if f.category == "chair")
        area_sf = room.area_sf
        area_m2 = round(area_sf * _SQFT_TO_SQM, 2) if area_sf is not None else None
        ws.append([
            "Floor 1", "GENERAL", space_type, subcategory, room.id, room.label or "",
            headcount,
            area_m2,
            round(area_sf, 1) if area_sf is not None else None,
            "Floor finish (assumed)", "Ceiling (assumed)",
            _furniture_elements(items),
        ])


def _general_sheet(ws: Worksheet) -> None:
    _header(ws, [
        "Floor Height", "Ceiling Height", "Door Height",
        "Glass Partition Height", "Glass Plaster Wall Height",
    ])
    ws.append([12, 9, 7, 8, 8])
    ws.append(["ft", "ft", "ft", "ft", "ft"])


def _main_summary_sheet(ws: Worksheet, layout: ExtractedLayout) -> None:
    _header(ws, [
        "Material Category", "Material Name", "Material ID", "Unit Type",
        "Quantity", "Unit Price", "Total cost",
    ])
    box_area = round(_box_area_sqft(layout), 1)
    rows: list[tuple[str, str, str, str, float]] = []

    for wall_type, length in sorted(_wall_length_by_type(layout).items()):
        rows.append(("Wall", f"{wall_type} wall", f"WALL-{wall_type.upper()}", "ft", round(length, 1)))
    rows.append(("Glass Partition", "Glass partition", "GLASS-PART", "ft", round(_glass_length(layout), 1)))
    rows.append(("Door", "Door", "DOOR", "number", float(len(layout.doors))))
    rows.append(("Floor", "Floor finish (assumed)", "FLOOR", "sqft", box_area))
    rows.append(("Ceiling", "Ceiling (assumed)", "CEILING", "sqft", box_area))

    for category, name, mat_id, unit, qty in rows:
        ws.append([category, name, mat_id, unit, qty, 0, 0])


def _bom_sheet(ws: Worksheet, rows: list[tuple[str, float, str]]) -> None:
    _header(ws, ["Material Name", "Quantity", "Unit", "Unit Price", "Total Price"])
    for name, qty, unit in rows:
        ws.append([name, round(qty, 1), unit, 0, 0])


def build_layout_takeoff(layout: ExtractedLayout) -> Workbook:
    """Build the 9-sheet Qbiq-grade takeoff workbook from a real ExtractedLayout. Pure."""
    assignment, most_unpolygoned = _assign_rooms(layout)

    wb = Workbook()
    _furniture_inventory_sheet(wb.active, layout, assignment)
    wb.active.title = "Furniture Inventory"
    _furniture_summary_sheet(wb.create_sheet("Furniture Inventory Summary"), layout)
    _inventory_sheet(wb.create_sheet("Inventory"), layout, assignment)
    _general_sheet(wb.create_sheet("General"))
    _main_summary_sheet(wb.create_sheet("Main Summary"), layout)

    box_area = round(_box_area_sqft(layout), 1)
    _bom_sheet(
        wb.create_sheet("BOM - Walls"),
        [(f"{t} wall", length, "ft") for t, length in sorted(_wall_length_by_type(layout).items())],
    )
    _bom_sheet(wb.create_sheet("BOM - Glass Partitions"),
               [("Glass partition", _glass_length(layout), "ft")])
    _bom_sheet(wb.create_sheet("BOM - Doors"), [("Door", float(len(layout.doors)), "number")])
    _bom_sheet(wb.create_sheet("BOM - Floors"), [("Floor finish (assumed)", box_area, "sqft")])
    _bom_sheet(wb.create_sheet("BOM - Ceilings"), [("Ceiling (assumed)", box_area, "sqft")])

    if most_unpolygoned:
        ws = wb.active
        ws.append([])
        ws.append([
            "Note: most rooms have no closed polygon — those items are Room ID '—' (Unassigned)."
        ])
    return wb
