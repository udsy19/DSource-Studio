"""Quantity-takeoff workbook — sheets, priced lines, provenance flags, derived shell figures."""

import pytest
from openpyxl import load_workbook
import io
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.database import Base
from app.models import Product
from app.floorplan.dxf_ingest import PlanModel
from app.takeoff.service import build_takeoff_workbook
from app.testfit.layout import FurnitureInstance, TestFit


@pytest.fixture
def db():
    engine = create_engine("sqlite://", connect_args={"check_same_thread": False})
    Base.metadata.create_all(bind=engine)
    s = sessionmaker(bind=engine, autoflush=False)()
    try:
        yield s
    finally:
        s.close()


@pytest.fixture
def seeded(db):
    products = [
        Product(manufacturer_code="HM", sku="AER1", name="Aeron Task Chair",
                list_price=1726.0, source="pricebook"),
        Product(manufacturer_code="SC", sku="SC-OLOGY-RECT", name="Ology Desk",
                list_price=500.0, source="sif"),
        Product(manufacturer_code="HM", sku="HM-RENEW-SS", name="Renew Sit-to-Stand",
                list_price=800.0, source="sif"),
        Product(manufacturer_code="HM", sku="HM-EVERYWHERE-6", name="Everywhere Table",
                list_price=2000.0, source="sif"),
    ]
    db.add_all(products)
    db.flush()
    return db


@pytest.fixture
def plan():
    return PlanModel(
        units="feet", sqft_factor=1.0,
        boundary=[(0, 0), (100, 0), (100, 100), (0, 100)],
        gross_area_sf=10000.0, core_area_sf=100.0, usable_area_sf=1000.0,
        columns=[], cores=[[(0, 0), (10, 0), (10, 10), (0, 10)]],
    )


@pytest.fixture
def fit():
    return TestFit(
        workstation_count=2, office_count=1, meeting_count=1, collab_count=0,
        instances=[
            FurnitureInstance("workstation", 0, 0, 6, 5),
            FurnitureInstance("workstation", 10, 0, 6, 5),
            FurnitureInstance("private_office", 20, 0, 12, 10),
            FurnitureInstance("meeting_room", 40, 0, 6, 5),  # area 30 -> 2 seats
        ],
    )


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


def test_three_sheets_exist(seeded, plan, fit):
    wb = build_takeoff_workbook(seeded, plan, fit)
    assert wb.sheetnames == ["Furniture Inventory", "Summary", "Areas & Shell"]


def test_inventory_rows_and_pricing_flags(seeded, plan, fit):
    wb = build_takeoff_workbook(seeded, plan, fit)
    rows = _rows(wb["Furniture Inventory"])

    assert rows[0] == ("Room ID", "Room Type", "Item", "Supplier", "Quantity",
                       "Unit Price (INR)", "Total (INR)", "Pricing")
    body = rows[1:]
    # 2 ws (2 lines each) + 1 office (2) + 1 meeting (2) = 8 lines
    assert len(body) == 8

    # The meeting room (room id 3) yields the table (est.) + 2 chairs (real).
    meeting = [r for r in body if r[0] == 3]
    table = next(r for r in meeting if r[2] == "Everywhere Table")
    chair = next(r for r in meeting if r[2] == "Aeron Task Chair")
    assert table[1] == "Meeting Room" and table[7] == "est."
    assert chair[4] == 2 and chair[6] == 3452.0 and chair[7] == "real"

    # Provenance: AER1 is the only price-book SKU -> 4 real lines, 4 est.
    assert sum(1 for r in body if r[7] == "real") == 4
    assert sum(1 for r in body if r[7] == "est.") == 4


def test_summary_totals(seeded, plan, fit):
    wb = build_takeoff_workbook(seeded, plan, fit)
    summary = dict(r for r in _rows(wb["Summary"])[1:])

    assert summary["Workstations"] == 2
    assert summary["Private offices"] == 1
    assert summary["Meeting rooms"] == 1
    assert summary["Collaboration zones"] == 0
    assert summary["Usable area (sf)"] == 1000.0
    assert summary["Density (usable sf / seat)"] == 333.3  # 1000 / (2+1) seats
    assert summary["Grand total (INR)"] == 12430.0
    assert summary["Real-price lines"] == 4
    assert summary["Estimated lines"] == 4


def test_areas_and_shell_perimeters(seeded, plan, fit):
    wb = build_takeoff_workbook(seeded, plan, fit)
    rows = dict(r for r in _rows(wb["Areas & Shell"]) if r and r[0])

    assert rows["Boundary perimeter (LF)"] == 400.0
    assert rows["Total core perimeter (LF)"] == 40.0
    assert rows["Gross area (sf)"] == 10000.0


def test_workbook_serializes_to_xlsx_bytes(seeded, plan, fit):
    wb = build_takeoff_workbook(seeded, plan, fit)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    assert load_workbook(buf).sheetnames[0] == "Furniture Inventory"
