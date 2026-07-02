"""Qbiq-grade layout takeoff — 9 sheets, summary aggregation, BOM lengths/counts from geometry."""

import io

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.ingestion.schema import Door, ExtractedLayout, FurnitureItem, Room, Wall
from app.main import app
from app.takeoff.layout_takeoff import build_layout_takeoff, build_program_summary

_XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _layout() -> ExtractedLayout:
    # Two rooms with closed polygons: A = [0,20]x[0,10], B = [20,40]x[0,10].
    room_a = [(0, 0), (20, 0), (20, 10), (0, 10), (0, 0)]
    room_b = [(20, 0), (40, 0), (40, 10), (20, 10), (20, 0)]
    return ExtractedLayout(
        source="cad",
        units="ft",
        bounds=(0.0, 0.0, 40.0, 10.0),  # box area = 400 sqft
        walls=[
            # drywall: two 20-ft segments = 40 ft total
            Wall(points=[(0, 0), (20, 0)], type="drywall"),
            Wall(points=[(20, 0), (40, 0)], type="drywall"),
            # glass: one 10-ft segment
            Wall(points=[(0, 10), (0, 0)], type="glass"),
        ],
        doors=[
            Door(x=5, y=0, width=3, rotation=0),
            Door(x=25, y=0, width=3, rotation=0),
        ],
        rooms=[
            Room(id="R-1", label="OFFICE 1", area_sf=200.0, polygon=room_a,
                 center=(10, 5), type="office"),
            Room(id="R-2", label="CONFERENCE", area_sf=200.0, polygon=room_b,
                 center=(30, 5), type="meeting"),
        ],
        furniture=[
            # Room A: a chair and a desk
            FurnitureItem(category="chair", block_name="Task Chair", brand=None, model=None,
                          x=2, y=2, w=2, h=2, rotation=0),
            FurnitureItem(category="desk", block_name="Desk", brand=None, model=None,
                          x=6, y=2, w=5, h=2.5, rotation=0),
            # Room B: two chairs and a table
            FurnitureItem(category="chair", block_name="Task Chair", brand=None, model=None,
                          x=22, y=2, w=2, h=2, rotation=0),
            FurnitureItem(category="chair", block_name="Task Chair", brand=None, model=None,
                          x=26, y=2, w=2, h=2, rotation=0),
            FurnitureItem(category="table", block_name="Conf Table", brand=None, model=None,
                          x=30, y=3, w=6, h=4, rotation=0),
            # A glass panel (counts toward glass length; max(w,h)=5) and a mullion (excluded)
            FurnitureItem(category="panel", block_name="Glazed", brand=None, model=None,
                          x=20, y=0, w=0.2, h=5, rotation=0),
            FurnitureItem(category="mullion", block_name="Mullion", brand=None, model=None,
                          x=20, y=0, w=0.1, h=5, rotation=0),
        ],
    )


def _rows(ws):
    return list(ws.iter_rows(values_only=True))


def test_all_nine_sheets():
    wb = build_layout_takeoff(_layout())
    assert wb.sheetnames == [
        "Furniture Inventory", "Furniture Inventory Summary", "Inventory", "General",
        "Main Summary", "BOM - Walls", "BOM - Glass Partitions", "BOM - Doors",
        "BOM - Floors", "BOM - Ceilings",
    ]


def test_furniture_inventory_descriptions_and_room_assignment():
    wb = build_layout_takeoff(_layout())
    rows = _rows(wb["Furniture Inventory"])
    assert rows[0] == ("Cost Code", "Floor", "Room ID", "Room Type", "Item Description",
                       "Supplier", "Quantity", "Unit Price", "Total Price")
    body = rows[1:]
    # 6 takeoff items (mullion excluded)
    assert len(body) == 6
    # First row carries "Can be customized" in Cost Code + Supplier; later rows blank.
    assert body[0][0] == "Can be customized" and body[0][5] == "Can be customized"
    assert body[1][0] == "" and body[1][5] == ""
    # Chair sized 2x2 ft -> W61 X L61 cm (2*30.48 = 60.96 -> 61).
    chair = next(r for r in body if r[4].startswith("Chairs"))
    assert chair[4] == "Chairs W61 X L61"
    # Point-in-polygon: the desk center (8.5, 3.25) is in Room A.
    desk = next(r for r in body if r[4].startswith("Desk Table"))
    assert desk[2] == "R-1"
    # The table center (33, 5) is in Room B.
    table = next(r for r in body if r[4].startswith("Table W"))
    assert table[2] == "R-2"


def test_summary_aggregation():
    wb = build_layout_takeoff(_layout())
    rows = _rows(wb["Furniture Inventory Summary"])
    assert rows[0] == ("Cost Code", "Item Description", "Supplier", "Quantity",
                       "Unit Price", "Total Price")
    agg = {r[1]: r[3] for r in rows[1:]}
    # 3 chairs aggregate to one line of qty 3.
    assert agg["Chairs W61 X L61"] == 3
    assert agg["Desk Table W152 X L76"] == 1  # 5ft->152, 2.5ft->76
    # Sorted by Item Description.
    descs = [r[1] for r in rows[1:]]
    assert descs == sorted(descs)


def test_inventory_headcount_and_elements():
    wb = build_layout_takeoff(_layout())
    rows = _rows(wb["Inventory"])
    assert rows[0][0] == "Floor" and rows[0][6] == "Headcount"
    by_room = {r[4]: r for r in rows[1:]}
    # Room B has 2 chairs -> headcount 2; department GENERAL.
    assert by_room["R-2"][1] == "GENERAL"
    assert by_room["R-2"][6] == 2
    assert by_room["R-1"][6] == 1
    # Furniture Elements lists per-room item counts.
    assert "Chairs: 2" in by_room["R-2"][11]
    # Assumed materials are flagged.
    assert by_room["R-1"][9] == "Floor finish (assumed)"


def test_bom_walls_length_by_type():
    wb = build_layout_takeoff(_layout())
    rows = {r[0]: r for r in _rows(wb["BOM - Walls"])[1:]}
    assert rows["drywall wall"][1] == 40.0  # two 20-ft segments
    assert rows["drywall wall"][2] == "ft"
    assert rows["glass wall"][1] == 10.0


def test_bom_glass_doors_floors():
    wb = build_layout_takeoff(_layout())
    glass = _rows(wb["BOM - Glass Partitions"])[1]
    assert glass[0] == "Glass partition" and glass[1] == 5.0  # max(0.2,5) of the one panel
    doors = _rows(wb["BOM - Doors"])[1]
    assert doors[1] == 2.0 and doors[2] == "number"
    floors = _rows(wb["BOM - Floors"])[1]
    assert floors[1] == 400.0 and floors[2] == "sqft"  # 40 x 10 box


def test_general_heights_defaults():
    wb = build_layout_takeoff(_layout())
    rows = _rows(wb["General"])
    assert rows[0][1] == "Ceiling Height"
    # value row: ceiling 9, door 7, glass partition 8
    assert rows[1] == (12, 9, 7, 8, 8)
    assert rows[2] == ("ft", "ft", "ft", "ft", "ft")


def test_main_summary_one_row_per_material():
    wb = build_layout_takeoff(_layout())
    rows = _rows(wb["Main Summary"])
    assert rows[0][0] == "Material Category"
    names = {r[1]: r for r in rows[1:]}
    assert names["drywall wall"][4] == 40.0
    assert names["Glass partition"][4] == 5.0
    assert names["Door"][4] == 2.0
    assert names["Floor finish (assumed)"][4] == 400.0


def test_serializes_to_xlsx_bytes():
    wb = build_layout_takeoff(_layout())
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    assert len(load_workbook(buf).sheetnames) == 10


# A layout shaped exactly as frontend/src/fitToLayout.ts emits it: source="generated", a
# perimeter wall, one synthesized room, and furniture WITHOUT room_id.
def _fit_shaped_dict() -> dict:
    return {
        "source": "generated",
        "units": "ft",
        "bounds": [0.0, 0.0, 20.0, 10.0],
        "walls": [
            {"type": "perimeter", "points": [[0, 0], [20, 0], [20, 10], [0, 10], [0, 0]]},
        ],
        "doors": [],
        "rooms": [
            {
                "id": "r0",
                "label": "Office",
                "area_sf": 200,
                "polygon": [[0, 0], [20, 0], [20, 10], [0, 10], [0, 0]],
                "center": [10, 5],
                "type": "office",
            },
        ],
        "furniture": [
            {"category": "workstation", "block_name": "", "brand": "", "model": "",
             "x": 2, "y": 2, "w": 5, "h": 2.5, "rotation": 0},
            {"category": "chair", "block_name": "", "brand": "", "model": "",
             "x": 8, "y": 2, "w": 2, "h": 2, "rotation": 0},
        ],
    }


def test_schema_parses_fit_shaped_layout_without_room_id():
    layout = ExtractedLayout.model_validate(_fit_shaped_dict())
    assert layout.source == "generated"
    assert all(f.room_id is None for f in layout.furniture)


def test_layout_takeoff_endpoint_streams_xlsx():
    client = TestClient(app)
    res = client.post("/api/layout/takeoff", json=_fit_shaped_dict())
    assert res.status_code == 200
    assert res.headers["content-type"] == _XLSX_MEDIA
    assert res.content
    assert len(load_workbook(io.BytesIO(res.content)).sheetnames) == 10


def test_program_summary_groups_by_family_with_areas():
    rows = _rows(build_program_summary(_layout())["Program Summary"])
    assert rows[0] == ("Family", "Room Type", "Count", "Total Area (sqf)",
                       "Total Area (m2)", "Avg Area (sqf)")
    body = {r[1]: r for r in rows[1:] if r[1]}
    # office room -> Offices family, 1 room, 200 sqf measured.
    assert body["office"][0] == "Offices"
    assert body["office"][2] == 1
    assert body["office"][3] == 200.0
    # meeting -> Conference; Offices sorts before Conference (family order).
    assert body["meeting"][0] == "Conference"
    families_in_order = [r[0] for r in rows[1:] if r[1]]
    assert families_in_order == ["Offices", "Conference"]


def test_program_summary_total_row():
    rows = _rows(build_program_summary(_layout())["Program Summary"])
    total = next(r for r in rows if r[0] == "Total")
    assert total[2] == 2  # two rooms
    assert total[3] == 400.0  # 200 + 200 sqf


def test_program_summary_endpoint_streams_xlsx():
    client = TestClient(app)
    res = client.post("/api/layout/program-summary", json=_fit_shaped_dict())
    assert res.status_code == 200
    assert res.headers["content-type"] == _XLSX_MEDIA
    assert load_workbook(io.BytesIO(res.content)).sheetnames == ["Program Summary"]
